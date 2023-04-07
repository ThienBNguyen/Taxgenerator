from pdf2image import convert_from_path
from threading import Lock
import cv2
import numpy as np
import re
import pytesseract
import openpyxl as xl
# Convert PDF to image using pdf2image
lock = Lock()
def convert_pdf(pdf_path, save_dir, res = 400):
    pages = convert_from_path(pdf_path, res )
    name_with_extension = pdf_path.rsplit('/')[-1]
    name = name_with_extension.rsplit('.')[0]
    # save all of the pages in pdf to images 
    # for idx, page in enumerate(pages):
    #     print(page)
    #     page.save(f'{save_dir}/{name}_{idx}.png', 'PNG')
    pages[0].save(f'{save_dir}/{name}.jpg', 'JPEG')
# run this function to convert pdf to img
with lock:
    convert_pdf(
    '/Users/thiennguyen/Google Drive/GitHub Project/TaxGemerator/readpdf/w2.pdf', '/Users/thiennguyen/Google Drive/GitHub Project/TaxGemerator/readpdf')
with lock:
    img = cv2.imread('./w2.jpg')
text = pytesseract.image_to_string(img)
employer_name = ""
employer_name_idx = 0
employee_name = ""
employee_name_idx = 0
employer_id = ""
employer_id_idx = 0
employer_ssa = ""
employee_wage = ""
employee_wage_idx = 0
federal_income_withheld = ""
social_security = ""
social_security_idx = 0
social_security_tax_withheld = ""
medicare_wages = ""
medicare_wages_idx = 0
medicare_wages_withheld = ""
# print(type(text)) return string
# Split the string by newline character to get a list of lines
lines = text.split('\n')
for idx, line in enumerate(lines):
    if line.find("Employer’s name") != -1:
        employer_name_idx = idx
        break
for idx, line in enumerate(lines):
    if line.find("Employee’s name") != -1:
        employee_name_idx = idx
        break
for idx, line in enumerate(lines):
    if line.find("Employee’s name") != -1:
        employee_name_idx = idx
        break
for idx, line in enumerate(lines):
    if line.find("FED ID number") != -1:
        employer_id_idx = idx
        break
for idx, line in enumerate(lines):
    if line.find("Wages, tips") != -1:
        employee_wage_idx = idx
        break
for idx, line in enumerate(lines):
    if line.find("Social security") != -1:
        social_security_idx = idx
        break
for idx, line in enumerate(lines):
    if line.find("Medicare wages") != -1:
        medicare_wages_idx = idx
        break
# Extract the second line (which contains the name)
employer_name = lines[employer_name_idx + 1] + " " + lines[employer_name_idx + 2] + " " + lines[employer_name_idx + 3]
employee_name = lines[employee_name_idx + 1] + " " + lines[employee_name_idx + 2] + " " + lines[employee_name_idx + 3]

employer_id_str = lines[employer_id_idx + 2].split(' ')
employer_id = employer_id_str[0]
employer_ssa = employer_id_str[1]
employee_wage_str = lines[employee_wage_idx + 1].split(' ')
employee_wage = employee_wage_str[0]
federal_income_withheld = employee_wage_str[1]
social_security_str = lines[employee_wage_idx + 1].split(' ')
social_security = social_security_str[0]
social_security_tax_withheld = social_security_str[1]
medicare_wages_str = lines[employee_wage_idx + 1].split(' ')
medicare_wages = medicare_wages_str[0]
medicare_wages_withheld = medicare_wages_str[1]
# match = re.search(r"\b\d{2}-\d{7}\b", employer_id_split)
# if match:
#     employer_id = match.group(0)
#     print(employer_id)

wb = xl.load_workbook('./Form 1040 Leadsheet .xlsx')
# print(wb.sheetnames)
W2_Generate = wb.create_sheet(title='W-2Generate', index=0)
ws_1040 = wb["1040"]
ws_w2 = wb["W-2"]
cell_taxpayer = ws_1040.cell(2, 5)
cell_taxpayer.value = employee_name
cell_ssn = ws_1040.cell(2, 10)
cell_ssn.value = employer_ssa
cell_w2_wages = ws_w2.cell(3, 5)
cell_w2_federal = ws_w2.cell(3, 6)
cell_w2_federal.value = float(federal_income_withheld)
cell_w2_wages.value = float(employee_wage)
# print(wb.sheetnames)
# wb.remove(wb['Sheet'])
# print(wb.sheetnames)
# W2_Generate = wb['W-2Generate']
W2_Generate['A1'] = "Employer’s name"
W2_Generate['B1'] = 'Employee’s name'
W2_Generate['C1'] = 'Employer’s FED ID'
W2_Generate['D1'] = 'Employee’s SSA number'
W2_Generate['E1'] = 'Wages, tips,'
W2_Generate['F1'] = 'Federal income'
W2_Generate['G1'] = 'Social security'
W2_Generate['H1'] = 'Social security tax withheld'
W2_Generate['I1'] = 'Medicare wages'
W2_Generate['J1'] = 'tips Medicare tax withheld'
W2_Generate.append([employer_name, employee_name,
          employer_id, employer_ssa, employee_wage, federal_income_withheld, social_security, social_security_tax_withheld, medicare_wages, medicare_wages_withheld])
# Saving the changes
wb.save('./Form 1040 Leadsheet .xlsx')
print('completed')
# print(name_line)
# Remove leading/trailing white space from the name line
# name_line = name_line.strip()
# print(name_line)
# Use regular expressions to extract the name
# name_match = re.match(r'^([A-Z ]+)$', name_line)

